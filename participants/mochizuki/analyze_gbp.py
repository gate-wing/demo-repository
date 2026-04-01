import pandas as pd
import re
import os
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 東京23区のリスト
WARDS_23 = [
    '千代田区', '中央区', '港区', '新宿区', '文京区', '台東区',
    '墨田区', '江東区', '品川区', '目黒区', '大田区', '世田谷区',
    '渋谷区', '中野区', '杉並区', '豊島区', '北区', '荒川区',
    '板橋区', '練馬区', '足立区', '葛飾区', '江戸川区'
]


def extract_area(address):
    """住所からエリア（都道府県）を抽出する"""
    if not address or pd.isna(address):
        return '不明'

    address = str(address)

    match = re.search(r'(東京都|北海道|(?:大阪|京都)府|\S{2,3}県)', address)
    if not match:
        return '不明'

    pref = match.group(1)

    if pref == '東京都':
        for ward in WARDS_23:
            if ward in address:
                return '東京都（23区）'
        return '東京都（23区以外）'

    return pref


def load_csv(filepath):
    """CSVファイルを読み込む（2行目の説明行をスキップ）"""
    for encoding in ['cp932', 'utf-8-sig', 'utf-8']:
        try:
            df = pd.read_csv(filepath, encoding=encoding, skiprows=[1], dtype=str)
            return df
        except UnicodeDecodeError:
            continue
    raise ValueError(f'文字コードが判別できません: {filepath}')


def main():
    data_dir = Path('data')
    output_file = 'analysis_result.xlsx'

    if not data_dir.exists():
        print('「data」フォルダが見つかりません。dataフォルダを作成してCSVファイルを入れてください。')
        return

    csv_files = sorted(data_dir.glob('*.csv'))
    if not csv_files:
        print('dataフォルダにCSVファイルがありません。')
        return

    all_data = []

    for csv_file in csv_files:
        month = csv_file.stem
        if month.endswith('.csv'):
            month = month[:-4]
        print(f'{month} を読み込み中...')

        try:
            df = load_csv(csv_file)
            df['月'] = month
            all_data.append(df)
        except Exception as e:
            print(f'{csv_file.name} の読み込みに失敗: {e}')

    if not all_data:
        print('読み込めるデータがありませんでした。')
        return

    combined = pd.concat(all_data, ignore_index=True)

    num_cols = [
        'Google 検索 - モバイル', 'Google 検索 - パソコン',
        'Google マップ - モバイル', 'Google マップ - パソコン',
        '通話', 'メッセージ', '予約', 'ルート', 'ウェブサイトのクリック',
        '料理の注文', 'フードメニューのクリック'
    ]

    for col in num_cols:
        if col in combined.columns:
            combined[col] = pd.to_numeric(combined[col], errors='coerce').fillna(0)

    combined['表示回数'] = (
        combined.get('Google 検索 - モバイル', 0) +
        combined.get('Google 検索 - パソコン', 0) +
        combined.get('Google マップ - モバイル', 0) +
        combined.get('Google マップ - パソコン', 0)
    )

    combined['アクション数'] = (
        combined.get('通話', 0) +
        combined.get('メッセージ', 0) +
        combined.get('予約', 0) +
        combined.get('ルート', 0) +
        combined.get('ウェブサイトのクリック', 0) +
        combined.get('料理の注文', 0) +
        combined.get('フードメニューのクリック', 0)
    )

    combined['エリア'] = combined['住所'].apply(extract_area)

    combined['アクション率(%)'] = (
        combined['アクション数'] / combined['表示回数'].replace(0, float('nan')) * 100
    ).round(2)

    # 月別集計
    monthly = combined.groupby('月').agg(
        表示回数合計=('表示回数', 'sum'),
        表示回数平均=('表示回数', 'mean'),
        表示回数中央値=('表示回数', 'median'),
        アクション数合計=('アクション数', 'sum'),
        アクション数平均=('アクション数', 'mean'),
        アクション数中央値=('アクション数', 'median'),
        店舗数=('ビジネス名', 'count')
    ).reset_index()
    monthly['表示回数平均'] = monthly['表示回数平均'].round(1)
    monthly['表示回数中央値'] = monthly['表示回数中央値'].round(1)
    monthly['アクション数平均'] = monthly['アクション数平均'].round(1)
    monthly['アクション数中央値'] = monthly['アクション数中央値'].round(1)
    monthly['アクション率(%)'] = (
        monthly['アクション数合計'] / monthly['表示回数合計'] * 100
    ).round(2)

    # エリア別集計
    area = combined.groupby('エリア').agg(
        表示回数合計=('表示回数', 'sum'),
        表示回数平均=('表示回数', 'mean'),
        表示回数中央値=('表示回数', 'median'),
        アクション数合計=('アクション数', 'sum'),
        アクション数平均=('アクション数', 'mean'),
        アクション数中央値=('アクション数', 'median'),
        アクション率平均=('アクション率(%)', 'mean'),
        アクション率中央値=('アクション率(%)', 'median'),
        店舗数=('ビジネス名', 'count')
    ).reset_index().sort_values('表示回数合計', ascending=False)
    area['表示回数平均'] = area['表示回数平均'].round(1)
    area['表示回数中央値'] = area['表示回数中央値'].round(1)
    area['アクション数平均'] = area['アクション数平均'].round(1)
    area['アクション数中央値'] = area['アクション数中央値'].round(1)
    area['アクション率平均'] = area['アクション率平均'].round(2)
    area['アクション率中央値'] = area['アクション率中央値'].round(2)
    area['アクション率(%)'] = (
        area['アクション数合計'] / area['表示回数合計'] * 100
    ).round(2)

    # エリア×月別集計
    area_monthly = combined.groupby(['エリア', '月']).agg(
        表示回数合計=('表示回数', 'sum'),
        アクション数合計=('アクション数', 'sum')
    ).reset_index()
    area_monthly['アクション率(%)'] = (
        area_monthly['アクション数合計'] / area_monthly['表示回数合計'] * 100
    ).round(2)

    # 相関分析
    correlation = combined[['表示回数', 'アクション数']].corr()

    # Excelに出力
    print(f'\n結果を {output_file} に出力中...')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        monthly.to_excel(writer, sheet_name='月別集計', index=False)
        area.to_excel(writer, sheet_name='エリア別集計', index=False)
        area_monthly.to_excel(writer, sheet_name='エリア×月別集計', index=False)
        correlation.to_excel(writer, sheet_name='相関分析')
        combined[['月', 'エリア', 'ビジネス名', '住所', '表示回数', 'アクション数', 'アクション率(%)']].to_excel(
            writer, sheet_name='全データ', index=False
        )

        # レポートサマリーシートを作成
        wb = writer.book
        ws = wb.create_sheet('レポートサマリー', 0)

        # スタイル定義
        title_font = Font(name='メイリオ', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
        body_font = Font(name='メイリオ', size=11)
        label_font = Font(name='メイリオ', size=11, bold=True)

        blue_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
        light_blue_fill = PatternFill(fill_type='solid', fgColor='BDD7EE')
        gray_fill = PatternFill(fill_type='solid', fgColor='595959')
        green_fill = PatternFill(fill_type='solid', fgColor='70AD47')
        orange_fill = PatternFill(fill_type='solid', fgColor='ED7D31')

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # タイトル
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Googleビジネスプロフィール 分析レポート'
        ws['A1'].font = title_font
        ws['A1'].fill = blue_fill
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 40

        # 分析期間
        ws.merge_cells('A2:F2')
        ws['A2'] = f'分析期間：{combined["月"].min()} 〜 {combined["月"].max()}　／　対象店舗数：{combined["ビジネス名"].nunique()} 店舗'
        ws['A2'].font = Font(name='メイリオ', size=11, color='FFFFFF')
        ws['A2'].fill = gray_fill
        ws['A2'].alignment = center
        ws.row_dimensions[2].height = 25

        ws.row_dimensions[3].height = 15

        # ── 全体サマリー ──
        ws.merge_cells('A4:F4')
        ws['A4'] = '■ 全体サマリー'
        ws['A4'].font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
        ws['A4'].fill = gray_fill
        ws['A4'].alignment = left
        ws.row_dimensions[4].height = 25

        total_display = int(combined['表示回数'].sum())
        total_action = int(combined['アクション数'].sum())
        total_rate = round(total_action / total_display * 100, 2) if total_display > 0 else 0
        best_month = monthly.loc[monthly['アクション率(%)'].idxmax(), '月']
        worst_month = monthly.loc[monthly['アクション率(%)'].idxmin(), '月']

        top_display_area_row = area.loc[area['表示回数平均'].idxmax()]
        top_display_area_name = top_display_area_row['エリア']
        top_display_area_avg = top_display_area_row['表示回数平均']
        top_display_area_med = top_display_area_row['表示回数中央値']

        top_action_area_row = area.loc[area['アクション率平均'].idxmax()]
        top_action_area_name = top_action_area_row['エリア']
        top_action_area_avg = top_action_area_row['アクション率平均']
        top_action_area_med = top_action_area_row['アクション率中央値']

        summary_data = [
            ('総表示回数', f'{total_display:,} 回'),
            ('総アクション数', f'{total_action:,} 回'),
            ('平均アクション率', f'{total_rate} %'),
            ('最もアクション率が高い月', best_month),
            ('最もアクション率が低い月', worst_month),
            ('表示回数トップエリア（1店舗平均）', f'{top_display_area_name}　平均: {top_display_area_avg:,.1f} 回 ／ 中央値: {top_display_area_med:,.1f} 回'),
            ('アクション率トップエリア（平均）', f'{top_action_area_name}　平均: {top_action_area_avg} % ／ 中央値: {top_action_area_med} %'),
        ]

        for i, (label, value) in enumerate(summary_data):
            row = 5 + i
            ws.merge_cells(f'A{row}:C{row}')
            ws.merge_cells(f'D{row}:F{row}')
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = light_blue_fill
            ws[f'A{row}'].alignment = left
            ws[f'D{row}'] = value
            ws[f'D{row}'].font = body_font
            ws[f'D{row}'].alignment = left
            ws.row_dimensions[row].height = 22

        ws.row_dimensions[5 + len(summary_data)].height = 15

        # ── 月別トレンド ──
        trend_start = 5 + len(summary_data) + 1
        ws.merge_cells(f'A{trend_start}:F{trend_start}')
        ws[f'A{trend_start}'] = '■ 月別トレンド'
        ws[f'A{trend_start}'].font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
        ws[f'A{trend_start}'].fill = gray_fill
        ws[f'A{trend_start}'].alignment = left
        ws.row_dimensions[trend_start].height = 25

        headers = ['月', '表示回数合計', '表示回数平均', '表示回数中央値', 'アクション数合計', 'アクション数平均', 'アクション数中央値', '店舗数', 'アクション率(%)']
        header_row = trend_start + 1
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_i, value=h)
            cell.font = Font(name='メイリオ', size=10, bold=True, color='FFFFFF')
            cell.fill = blue_fill
            cell.alignment = center
        ws.row_dimensions[header_row].height = 20

        for r_i, row_data in monthly.iterrows():
            data_row = header_row + 1 + r_i
            values = [
                row_data['月'], int(row_data['表示回数合計']), row_data['表示回数平均'], row_data['表示回数中央値'],
                int(row_data['アクション数合計']), row_data['アクション数平均'], row_data['アクション数中央値'],
                int(row_data['店舗数']), row_data['アクション率(%)']
            ]
            for col_i, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col_i, value=val)
                cell.font = body_font
                cell.alignment = center
                if r_i % 2 == 0:
                    cell.fill = PatternFill(fill_type='solid', fgColor='EBF3FB')
            ws.row_dimensions[data_row].height = 20

        area_start = header_row + 1 + len(monthly) + 2

        # ── エリア別ランキング ──
        ws.merge_cells(f'A{area_start}:J{area_start}')
        ws[f'A{area_start}'] = '■ エリア別ランキング（表示回数順）'
        ws[f'A{area_start}'].font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
        ws[f'A{area_start}'].fill = gray_fill
        ws[f'A{area_start}'].alignment = left
        ws.row_dimensions[area_start].height = 25

        area_headers = ['エリア', '店舗数', '表示回数合計', '表示回数平均', '表示回数中央値', 'アクション数合計', 'アクション数平均', 'アクション数中央値', 'アクション率平均(%)', 'アクション率中央値(%)']
        area_header_row = area_start + 1
        for col_i, h in enumerate(area_headers, 1):
            cell = ws.cell(row=area_header_row, column=col_i, value=h)
            cell.font = Font(name='メイリオ', size=10, bold=True, color='FFFFFF')
            cell.fill = blue_fill
            cell.alignment = center
        ws.row_dimensions[area_header_row].height = 20

        for r_i, row_data in area.reset_index(drop=True).iterrows():
            data_row = area_header_row + 1 + r_i
            values = [
                row_data['エリア'], int(row_data['店舗数']), int(row_data['表示回数合計']), row_data['表示回数平均'], row_data['表示回数中央値'],
                int(row_data['アクション数合計']), row_data['アクション数平均'], row_data['アクション数中央値'],
                row_data['アクション率平均'], row_data['アクション率中央値']
            ]
            for col_i, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col_i, value=val)
                cell.font = body_font
                cell.alignment = center
                if r_i % 2 == 0:
                    cell.fill = PatternFill(fill_type='solid', fgColor='EBF3FB')
            ws.row_dimensions[data_row].height = 20

        # 列幅を調整
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18

    print(f'\n完了！{output_file} を開いて結果を確認してください。')
    print(f'\n概要:')
    print(f'  分析期間: {combined["月"].min()} 〜 {combined["月"].max()}')
    print(f'  総店舗数: {combined["ビジネス名"].nunique()} 店舗')
    print(f'  エリア数: {combined["エリア"].nunique()} エリア')


if __name__ == '__main__':
    main()
